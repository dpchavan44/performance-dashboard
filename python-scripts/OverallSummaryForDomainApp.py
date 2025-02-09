import xlrd
import os
import xlwt
from openpyxl import load_workbook
import datetime

def CopyExcelsheet():
    # Join the current directory and the file name
    current_dir = os.getcwd()
    file_name = 'QualityPlanningSummaryData.xlsx'
    path = os.path.join(current_dir,'../files', file_name)
    
    # Print the current working directory and the full path
    print("Current Working Directory:", current_dir)
    print("Full Path:", path)
    
    # Check if the file exists
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return
    
    try:
        book = xlrd.open_workbook(path)
    except FileNotFoundError as e:
        print(f"Error: {e}")
        return
    wb = xlwt.Workbook()
    ws = wb.add_sheet('OverallSummary1',cell_overwrite_ok = True)
    numberOfModules = book.nsheets
    row = 0
    for sheet in book.sheets():
      moduleName =sheet.name
      print("*********************************************************")
      print(moduleName)
      ws.write(row, 0,moduleName )
      sheet = book.sheet_by_name(moduleName)
      
      totalNumberOfUseCase = sheet.nrows-2
      ws.write(row, 3,totalNumberOfUseCase )
      #last baseline date
      xldate = sheet.cell(0,sheet.ncols-2).value
      
      tempDate = datetime.datetime(1900, 1, 1)
      deltaDays = datetime.timedelta(days=float(xldate)-2)
      TheTime = (tempDate + deltaDays )
      newdate =  TheTime.strftime("%m/%d/%Y")
      ws.write(row, 1,newdate)
      if moduleName.__contains__('Workflow') or moduleName.__contains__('4GD-AW'):
        print("Do nothing ")
      else :
        urlData ="_"+moduleName
        ws.write(row, 8,urlData)
      #last baseline name
      baseline  = sheet.cell(1,sheet.ncols-2).value
      ws.write(row, 2,baseline)
      
      add = 0
      for row_index in range(1, sheet.nrows):
        #print(sheet.cell(row_index,sheet.ncols-1).value)
        comments = sheet.cell(row_index,sheet.ncols-1).value
        if comments.__contains__('New') or comments.__contains__('new') :
          add = add + 1
      #reading last baseline value
      listOfLatestBaselineValue = []
      for row_Index in range(2, sheet.nrows):
        #print(sheet.cell(row_Index,sheet.ncols-2).value)
        latestValue = sheet.cell(row_Index,sheet.ncols-2).value
        listOfLatestBaselineValue.append(latestValue)
      #read all baseline values (whole second row from exel)
      listofBaslines=[]  
      for col in range(sheet.ncols):
        data= sheet.cell_value(1,col)  
        listofBaslines.append(data)
      #print(listofBaslines)
      #print(any("RTM" in s for s in listofBaslines))
      #get all RTM baseline indexes
      allindextlistwithRTM=[]
      for idx, val in enumerate(listofBaslines):
        #print(idx, val)
        if val.__contains__("RTM") or val.__contains__("rtm") or val.__contains__("Rtm") or val.__contains__("Master"):
          #print(idx)
          allindextlistwithRTM.append(idx)
      #print(allindextlistwithRTM)
      #read last index of RTM baseline
      rtmcolumnIndex = int(allindextlistwithRTM[-1])
      listOfLatestRTMvalues= [] #get all RTM values
      for row_Index in range(2,sheet.nrows):
        #print(sheet.cell(row_Index,sheet.ncols-2).value)
        RTMlatestValue = sheet.cell(row_Index,rtmcolumnIndex).value
        listOfLatestRTMvalues.append(RTMlatestValue)
      #read release target values
      listOfReleaseTargetValue = []
      for row_Index in range(2,sheet.nrows):
      	ReleaseTargetValue =  sheet.cell(row_Index,4).value
      	listOfReleaseTargetValue.append(ReleaseTargetValue)
      
      #print(listOfReleaseTargetValue)
	    #read fimal target values
      listOfFinalTargetValue = []
      for row_Index in range(2,sheet.nrows):
      	FinalTargetValue =  sheet.cell(row_Index,3).value
      	listOfFinalTargetValue.append(FinalTargetValue)
      
      #print(listOfFinalTargetValue) 
	  
      for n ,i in enumerate(listOfReleaseTargetValue):
        if i =='Ignore':
          listOfReleaseTargetValue[n] = 0
		  
      for n ,i in enumerate(listOfFinalTargetValue):
        if i =='Ignore':
          listOfFinalTargetValue[n] = 0
		  
      for n ,i in enumerate(listOfLatestRTMvalues):
        if i =='Ignore':
          listOfLatestRTMvalues[n] = 0

      for n ,i in enumerate(listOfLatestBaselineValue):
        if i =='Ignore':
          listOfLatestBaselineValue[n] = 0

      '''print("listOfLatestBaselineValue :")
      print(listOfLatestBaselineValue)
      print("listOfLatestRTMvalues :")
      print(listOfLatestRTMvalues)
      print("listOfReleaseTargetValue :")
      print(listOfReleaseTargetValue)
      print("listOfFinalTargetValue :")
      print(listOfFinalTargetValue)'''
	  
      ImpreovedUseCase = []
      DegradedUseCase = []
      imporved = 0
      nutral = 0
      degraded = 0
      failed = 0
      for i in range(0,len(listOfLatestRTMvalues)):

        if  listOfLatestRTMvalues[i] > listOfLatestBaselineValue[i] :
          diff=listOfLatestRTMvalues[i] - listOfLatestBaselineValue[i]
          if (diff/listOfLatestRTMvalues[i]) > 0.1 and diff>=1  and listOfLatestBaselineValue[i] !=0:
            imporved = imporved +1
          else :
            nutral = nutral +1
        elif listOfLatestRTMvalues[i] < listOfLatestBaselineValue[i] and listOfLatestRTMvalues[i] !=0:
          diffforreg=  listOfLatestBaselineValue[i]- listOfLatestRTMvalues[i]
          if  (diffforreg/listOfLatestRTMvalues[i]) >= 0.1 and diffforreg >1:
            degraded = degraded +1
          else :
            nutral = nutral+1
        elif listOfLatestRTMvalues[i] == listOfLatestBaselineValue[i]:
          diff1 = 0
          nutral = nutral+1
        
        if isinstance(listOfLatestBaselineValue[i], float) and listOfLatestBaselineValue[i] == 0.0:
            failed = failed + 1
            
      allotherthanNew = imporved + nutral + degraded
      newUseCaseData = totalNumberOfUseCase - allotherthanNew
      #print(newUseCaseData)         
      #print(imporved)
      #print(nutral)
      #print(degraded)
      ws.write(row, 4,degraded)
      ws.write(row, 5,imporved)
      ws.write(row, 6,nutral)
      ws.write(row, 7,newUseCaseData)
      ws.write(row, 11, failed)

      Ndegraded = 0
      for i in range(0,len(listOfReleaseTargetValue)):

        if  listOfReleaseTargetValue[i] > listOfLatestBaselineValue[i] :
          diff=listOfReleaseTargetValue[i] - listOfLatestBaselineValue[i]
          if (diff/listOfReleaseTargetValue[i]) > 0.1 and diff>=1  and listOfLatestBaselineValue[i] !=0:
            imporved = imporved +1
          else :
            nutral = nutral +1
        elif listOfReleaseTargetValue[i] < listOfLatestBaselineValue[i] and listOfReleaseTargetValue[i] !=0:
          diffforreg=  listOfLatestBaselineValue[i]- listOfReleaseTargetValue[i]
          if  (diffforreg/listOfReleaseTargetValue[i]) > 0.1 and diffforreg >=1:
            Ndegraded = Ndegraded +1
          else :
            nutral = nutral+1
        elif listOfReleaseTargetValue[i] == listOfLatestBaselineValue[i]:
          diff1 = 0
          nutral = nutral+1
      #regressed from release target
      ws.write(row, 9,Ndegraded)

      Ndegraded = 0
      for i in range(0,len(listOfFinalTargetValue)):

        if  listOfFinalTargetValue[i] > listOfLatestBaselineValue[i] :
          diff=listOfFinalTargetValue[i] - listOfLatestBaselineValue[i]
          if (diff/listOfFinalTargetValue[i]) > 0.1 and diff>=1  and listOfLatestBaselineValue[i] !=0:
            imporved = imporved +1
          else :
            nutral = nutral +1
        elif listOfFinalTargetValue[i] < listOfLatestBaselineValue[i] and listOfFinalTargetValue[i] !=0:
          diffforreg=  listOfLatestBaselineValue[i]- listOfFinalTargetValue[i]
          if  (diffforreg/listOfFinalTargetValue[i]) > 0.1 and diffforreg >=1:
            Ndegraded = Ndegraded +1
          else :
            nutral = nutral+1

        elif listOfFinalTargetValue[i] == listOfLatestBaselineValue[i]:
          diff1 = 0
          nutral = nutral+1
      #regressed from final target
      ws.write(row, 10,Ndegraded)
	  
      row = row+1
	  
    #wb.save('D:\\dashboard\\OverallSummary\\New\\detailedReportDomainApp.xls')
    pathToSave = os.path.join(current_dir,'../files', 'detailedReportDomainApp.xls')
    wb.save(pathToSave)
    print("#############################################Updated Excel file")

CopyExcelsheet()